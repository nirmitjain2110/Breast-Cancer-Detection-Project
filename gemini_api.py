# import os
# import pandas as pd
# import google.generativeai as genai
# from openpyxl import Workbook, load_workbook

# # --- Configuration ---
# GOOGLE_API_KEY = "AIzaSyABvfj3hAaOotaLI7vV_tcy02ISq2KRQEM"
# genai.configure(api_key=GOOGLE_API_KEY)

# # Define input and output file names
# input_excel_file = "Modified_data.xlsx"
# output_excel_file = "features_gemini.xlsx"

# # --- Excel Writing Function (No changes needed) ---
# def writeInSheets(file_name, tumor_size, er, pr, her2):
#     """Writes the extracted features into an Excel spreadsheet."""
#     if not os.path.exists(output_excel_file):
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Features"
#         headers = ["Source File Name", "Tumor Size", "ER", "PR", "HER2"]
#         ws.append(headers)
#         wb.save(output_excel_file)

#     wb = load_workbook(output_excel_file)
#     ws = wb.active
#     ws.append([file_name, tumor_size, er, pr, her2])
#     wb.save(output_excel_file)

# # --- Modified Gemini QA Function ---
# def get_answer_from_gemini(question, context, model):
#     """
#     Gets an answer from Gemini based on the provided context.
#     """
#     prompt = f"""
#     Based ONLY on the context provided below, answer the question.
#     Provide only the direct value or answer, without any introductory phrases or explanations.

#     CONTEXT:
#     {context}

#     QUESTION:
#     {question}
#     """
#     try:
#         response = model.generate_content(prompt)
#         return response.text.strip()
#     except Exception as e:
#         print(f"An error occurred while calling Gemini API: {e}")
#         return "Error"

# # --- Main Execution ---

# # 1. Initialize the Gemini Model
# print("Initializing Gemini model...")
# gemini_model = genai.GenerativeModel('gemini-2.5-flash-lite')
# # 2. Load the input Excel file using pandas
# try:
#     print(f"Reading data from {input_excel_file}...")
#     df = pd.read_excel(input_excel_file)
# except FileNotFoundError:
#     print(f"Error: The input file '{input_excel_file}' was not found.")
#     exit()

# # Define the names of the columns in your input file
# # MODIFY THESE if your column names are different
# filename_column = 'File_name'
# report_text_column = 'ExtractedText'

# if report_text_column not in df.columns or filename_column not in df.columns:
#     print(f"Error: Make sure your Excel file has columns named '{filename_column}' and '{report_text_column}'.")
#     exit()

# # 3. Process each row in the Excel file
# print(f"Found {len(df)} rows to process.")
# for index, row in df.head(5).iterrows():
#     file_name = row[filename_column]
#     report_context = row[report_text_column]

#     print(f"\n--- Processing Report: {file_name} (Row {index + 2}) ---")

#     # Check if the report context is valid
#     if not isinstance(report_context, str) or len(report_context) < 10:
#         print("Skipping row due to empty or invalid report text.")
#         writeInSheets(file_name, "Skipped", "Skipped", "Skipped", "Skipped")
#         continue

#     # 4. Define questions and get answers for the current row
#     q1 = "What is the tumor size in cm?"
#     q2 = "Is Estrogen Receptor (ER) positive or negative?"
#     q3 = "Is Progesterone Receptor (PR) positive or negative?"
#     q4 = "What is the HER-2/neu status (e.g., negative, positive, 1+, 2+, 3+)?"

#     tumor_size = get_answer_from_gemini(q1, report_context, gemini_model)
#     er = get_answer_from_gemini(q2, report_context, gemini_model)
#     pr = get_answer_from_gemini(q3, report_context, gemini_model)
#     her2 = get_answer_from_gemini(q4, report_context, gemini_model)
    
#     print(f"  Tumor Size: {tumor_size}")
#     print(f"  ER Status: {er}")
#     print(f"  PR Status: {pr}")
#     print(f"  HER2 Status: {her2}")

#     # 5. Write the extracted results to the output Excel file
#     writeInSheets(file_name, tumor_size, er, pr, her2)

# print(f"\n✅ Processing complete. All data written to {output_excel_file}")


import os
import json
import pandas as pd
import google.generativeai as genai
from openpyxl import Workbook, load_workbook

# --- Configuration ---
GOOGLE_API_KEY = "AIzaSyBszzy2NBXgfUJfnDFf96q6b2XJbZBBUVs"   # replace with your key
genai.configure(api_key=GOOGLE_API_KEY)

# Define input and output file names
input_excel_file = "Modified_data.xlsx"
output_excel_file = "features_gemini.xlsx"

# --- Excel Writing Function ---
def writeInSheets(file_name, tumor_size, er, pr, her2):
    """Writes the extracted features into an Excel spreadsheet."""
    if not os.path.exists(output_excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Features"
        headers = ["Source File Name", "Tumor Size", "ER", "PR", "HER2"]
        ws.append(headers)
        wb.save(output_excel_file)

    wb = load_workbook(output_excel_file)
    ws = wb.active
    ws.append([file_name, tumor_size, er, pr, her2])
    wb.save(output_excel_file)

# --- Gemini Extraction Function ---
def extract_features_with_gemini(report_context, model):
    """
    Extracts tumor size, ER, PR, HER2 from pathology report using JSON schema prompting.
    """
    prompt = f"""
    You are analyzing a pathology report. 
    Extract the following features and normalize synonyms:

    - Tumor Size (in cm).
    - ER (Estrogen Receptor) status: "Positive" or "Negative".
        Synonyms: Estrogen receptor, ER receptor.
    - PR (Progesterone Receptor) status: "Positive" or "Negative".
        Synonyms: Progesterone receptor, PR receptor.
    - HER2/neu status: "Negative", "Positive", "1+", "2+", "3+".
        Synonyms: HER-2, HER2/neu, HER2 receptor, HER2 amplification.

    If not found, write "Not Reported".  

    Provide the result in strict JSON format with keys: 
    tumor_size, er, pr, her2.

    CONTEXT:
    {report_context}
    """
    try:
        response = model.generate_content(prompt)
        text_output = response.text.strip()

        # Try parsing JSON safely
        try:
            result = json.loads(text_output)
        except json.JSONDecodeError:
            # fallback if Gemini returns extra text
            text_output = text_output[text_output.find("{") : text_output.rfind("}")+1]
            result = json.loads(text_output)

        return (
            result.get("tumor_size", "Not Reported"),
            result.get("er", "Not Reported"),
            result.get("pr", "Not Reported"),
            result.get("her2", "Not Reported"),
        )

    except Exception as e:
        print(f"An error occurred while calling Gemini API: {e}")
        return ("Error", "Error", "Error", "Error")

# --- Main Execution ---
print("Initializing Gemini model...")
gemini_model = genai.GenerativeModel('gemini-2.5-flash-lite')

# Load input Excel
try:
    print(f"Reading data from {input_excel_file}...")
    df = pd.read_excel(input_excel_file)
except FileNotFoundError:
    print(f"Error: The input file '{input_excel_file}' was not found.")
    exit()

# Define expected columns
filename_column = 'File_name'
report_text_column = 'ExtractedText'

if report_text_column not in df.columns or filename_column not in df.columns:
    print(f"Error: Make sure your Excel file has columns named '{filename_column}' and '{report_text_column}'.")
    exit()

# Process rows
print(f"Found {len(df)} rows to process.")
for index, row in df.iterrows():
    file_name = row[filename_column]
    report_context = row[report_text_column]

    print(f"\n--- Processing Report: {file_name} (Row {index + 2}) ---")

    if not isinstance(report_context, str) or len(report_context) < 10:
        print("Skipping row due to empty or invalid report text.")
        writeInSheets(file_name, "Skipped", "Skipped", "Skipped", "Skipped")
        continue

    tumor_size, er, pr, her2 = extract_features_with_gemini(report_context, gemini_model)

    print(f"  Tumor Size: {tumor_size}")
    print(f"  ER Status: {er}")
    print(f"  PR Status: {pr}")
    print(f"  HER2 Status: {her2}")

    writeInSheets(file_name, tumor_size, er, pr, her2)

print(f"\n✅ Processing complete. All data written to {output_excel_file}")
