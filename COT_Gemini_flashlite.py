import os
import json
import pandas as pd
import google.generativeai as genai
from openpyxl import Workbook, load_workbook
from sklearn.metrics.pairwise import cosine_similarity
import numpy as np

# --- Configuration ---
GOOGLE_API_KEY = ""  
genai.configure(api_key=GOOGLE_API_KEY)

input_excel_file = "Modified_data.xlsx"
output_excel_file = "biomarkers_gemini_with_steps.xlsx"

# --- RAG Knowledge Base ---
BIOMARKER_KB = {
    "ER": "Estrogen receptor, hormone receptor positive cancers respond to hormone therapy.",
    "PR": "Progesterone receptor, hormone receptor linked with estrogen receptor.",
    "HER2": "HER2 or c-erbB2 amplification, indicates aggressive growth and targeted therapy eligibility.",
    "Ki-67": "Ki-67 is a proliferation index expressed as percentage, reflects tumor growth rate.",
    "p53": "TP53 mutation status, indicates tumor suppressor gene alteration, linked with poor prognosis.",
    "Tumor Size": "Size of the tumor given in cm, can appear as single or three-dimensional values like 2.3 x 1.5 x 1.2 cm."
}

# Create embeddings for RAG reference
embed_model = genai.embed_content(model="models/text-embedding-004", content=list(BIOMARKER_KB.values()))
kb_embeddings = np.array(embed_model["embedding"])

# --- Excel Writing Function ---
def writeInSheets(file_name, tumor_size, er, pr, her2, ki67, p53, rationale):
    """Writes the extracted features and short stepwise rationale into an Excel spreadsheet."""
    if not os.path.exists(output_excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Biomarkers"
        headers = ["Source File Name", "Tumor Size", "ER", "PR", "HER2", "Ki-67", "p53", "Extraction_Rationale"]
        ws.append(headers)
        wb.save(output_excel_file)

    wb = load_workbook(output_excel_file)
    ws = wb.active
    ws.append([file_name, tumor_size, er, pr, her2, ki67, p53, rationale])
    wb.save(output_excel_file)

# --- RAG Retrieval Function ---
def retrieve_context(report_text):
    """Finds the most relevant biomarker context using embeddings."""
    response = genai.embed_content(model="models/text-embedding-004", content=report_text)
    query_emb = np.array(response["embedding"])
    sims = cosine_similarity([query_emb], kb_embeddings)[0]
    top_idx = np.argmax(sims)
    return list(BIOMARKER_KB.values())[top_idx]

# --- Gemini Extraction Function (externally stepwise rationale) ---
def extract_features_with_gemini(report_context, model):
    """
    Extracts HER2/neu, ER, PR, Ki-67, p53, and tumor size from a breast cancer pathology report.
    Produces a JSON with the five fields plus an explicit external 'extraction_rationale' (short stepwise log).
    NOTE: This prompt requests an external, structured explanation — it does NOT attempt to elicit internal chain-of-thought.
    """

    rag_context = retrieve_context(report_context)

    prompt = f"""
You are a medical data extraction system. Do NOT reveal internal chain-of-thought or private model reasoning.
Instead, produce a concise, external step-by-step extraction log (1-3 short bullets per biomarker) called "extraction_rationale".

Task:
From the pathology report below, extract and return a JSON object with these keys:
- tumor_size (string; include units and laterality if present, or "Not Reported")
- her2 (one of: "Negative", "Equivocal", "Positive", "Ordered, result pending", or "Not Reported")
- er (one of: "Positive", "Negative", or "Not Reported"; if percent/score present include, e.g. "Positive (98%, 3+)")
- pr (same format as ER)
- ki67 (percentage + category: "12% (Borderline)" or "Not Reported")
- p53 (one of: "Positive", "Negative", "Not Reported")

Also include:
- extraction_rationale: a JSON object mapping each biomarker name to a short array of 1-3 succinct bullet strings that list the textual cues used to decide the value (e.g., ["ER: 'positive reaction 98% 3+' found", "Percent >=1% => Positive"]).

Reference rules for categorization (for model only; do not print these as part of rationale):
- ER/PR: Negative <1%, Positive ≥1%
- Ki-67: Low <10%, Borderline 10–20%, High >20%
- HER2: Negative 0–1+, Equivocal 2+, Positive 3+ (equivocal often reflexed to FISH)

If a field is not present in the report, set it to "Not Reported" and in rationale write a single bullet like "Not found".

Output format (strict JSON). Example:
{{
  "tumor_size": "6.0 x 4.0 x 3.5 cm (left)",
  "her2": "Ordered, result pending",
  "er": "Not Reported",
  "pr": "Not Reported",
  "ki67": "Not Reported",
  "p53": "Not Reported",
  "extraction_rationale": {{
      "tumor_size": ["Found 'forming a 6.0 x 4.0 x 3.5 cm tumor'"],
      "her2": ["Phrase 'Her-2/NEU has been ordered' -> Ordered, result pending"],
      "er": ["Not found"]
  }}
}}

RAG CONTEXT: {rag_context}

PATHOLOGY REPORT:
{report_context}

Return only valid JSON—no extra commentary.
"""

    try:
        response = model.generate_content(prompt)
        text_output = response.text.strip()

        # Try direct JSON parse
        try:
            result = json.loads(text_output)
        except json.JSONDecodeError:
            # Extract JSON substring heuristically
            start = text_output.find("{")
            end = text_output.rfind("}")
            if start != -1 and end != -1 and end > start:
                json_str = text_output[start:end+1]
                result = json.loads(json_str)
            else:
                raise

        # Normalize fields and rationale
        tumor_size = result.get("tumor_size", "Not Reported")
        er = result.get("er", "Not Reported")
        pr = result.get("pr", "Not Reported")
        her2 = result.get("her2", "Not Reported")
        ki67 = result.get("ki67", "Not Reported")
        p53 = result.get("p53", "Not Reported")

        # extraction_rationale as compact string for Excel (join bullets)
        rationale_obj = result.get("extraction_rationale", {})
        # Convert to a single string: each field -> joined bullets separated by ' | '
        rationale_parts = []
        for key in ["tumor_size","her2","er","pr","ki67","p53"]:
            bullets = rationale_obj.get(key, [])
            if isinstance(bullets, list):
                joined = " ; ".join(bullets) if bullets else "Not found"
            elif isinstance(bullets, str):
                joined = bullets
            else:
                joined = "Not found"
            rationale_parts.append(f"{key}: {joined}")
        rationale_str = " || ".join(rationale_parts)

        return tumor_size, er, pr, her2, ki67, p53, rationale_str

    except Exception as e:
        print(f"Error during extraction: {e}")
        # Return Not Reported and include an error rationale
        return "Not Reported", "Not Reported", "Not Reported", "Not Reported", "Not Reported", "Not Reported", f"error: {str(e)}"

# --- Main Execution ---
print("Initializing Gemini model...")
gemini_model = genai.GenerativeModel('gemini-2.5-flash-lite')

try:
    df = pd.read_excel(input_excel_file)
except FileNotFoundError:
    print(f"Error: '{input_excel_file}' not found.")
    exit()

filename_column = 'File_name'
report_text_column = 'ExtractedText'

if filename_column not in df.columns or report_text_column not in df.columns:
    print(f"Error: Excel file must contain '{filename_column}' and '{report_text_column}' columns.")
    exit()

print(f"Processing {len(df)} pathology reports...")

for index, row in df.iterrows():
    file_name = row[filename_column]
    report_text = row[report_text_column]

    if not isinstance(report_text, str) or len(report_text.strip()) < 10:
        writeInSheets(file_name, "Skipped", "Skipped", "Skipped", "Skipped", "Skipped", "Skipped", "Skipped")
        continue

    print(f"\n--- Processing Report: {file_name} ---")
    tumor_size, er, pr, her2, ki67, p53, rationale = extract_features_with_gemini(report_text, gemini_model)

    print(f"Tumor Size: {tumor_size}")
    print(f"ER: {er}")
    print(f"PR: {pr}")
    print(f"HER2: {her2}")
    print(f"Ki-67: {ki67}")
    print(f"p53: {p53}")
    print(f"Rationale: {rationale}")

    writeInSheets(file_name, tumor_size, er, pr, her2, ki67, p53, rationale)

print(f"\n✅ Extraction complete. Results saved to {output_excel_file}")
