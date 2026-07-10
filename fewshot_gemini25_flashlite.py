import os
import json
import pandas as pd
import google.generativeai as genai
from openpyxl import Workbook, load_workbook
from sklearn.metrics.pairwise import cosine_similarity
import numpy as np

# --- Configuration ---
GOOGLE_API_KEY = "AIzaSyBszzy2NBXgfUJfnDFf96q6b2XJbZBBUVs"  # Replace with your Gemini API key
genai.configure(api_key=GOOGLE_API_KEY)

input_excel_file = "Modified_data.xlsx"
output_excel_file = "biomarkers_gemini.xlsx"

# --- RAG Knowledge Base ---
BIOMARKER_KB = {
    "ER": "Estrogen receptor, hormone receptor positive cancers respond to hormone therapy.",
    "PR": "Progesterone receptor, hormone receptor linked with estrogen receptor.",
    "HER2": "HER2 or c-erbB2 amplification, indicates aggressive growth and targeted therapy eligibility.",
    "Ki-67": "Ki-67 is a proliferation index expressed as percentage, reflects tumor growth rate.",
    "p53": "TP53 mutation status, indicates tumor suppressor gene alteration, linked with poor prognosis.",
    "Tumor Size": "Size of the tumor given in cm, can appear as single or three-dimensional values like 2.3 x 1.5 x 1.2 cm."
}

# Create embeddings for RAG reference
embed_model = genai.embed_content(model="models/text-embedding-004", content=list(BIOMARKER_KB.values()))
kb_embeddings = np.array(embed_model["embedding"])

# --- Excel Writing Function ---
def writeInSheets(file_name, tumor_size, er, pr, her2, ki67, p53):
    """Writes the extracted features into an Excel spreadsheet."""
    if not os.path.exists(output_excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Biomarkers"
        headers = ["Source File Name", "Tumor Size", "ER", "PR", "HER2", "Ki-67", "p53"]
        ws.append(headers)
        wb.save(output_excel_file)

    wb = load_workbook(output_excel_file)
    ws = wb.active
    ws.append([file_name, tumor_size, er, pr, her2, ki67, p53])
    wb.save(output_excel_file)

# --- RAG Retrieval Function ---
def retrieve_context(report_text):
    """Finds the most relevant biomarker context using embeddings."""
    response = genai.embed_content(model="models/text-embedding-004", content=report_text)
    query_emb = np.array(response["embedding"])
    sims = cosine_similarity([query_emb], kb_embeddings)[0]
    top_idx = np.argmax(sims)
    return list(BIOMARKER_KB.values())[top_idx]

# --- Gemini Extraction Function ---
def extract_features_with_gemini(report_context, model):
    """
    Extracts HER2/neu, ER, PR, Ki-67, p53, and tumor size from a breast cancer pathology report.
    Returns structured JSON output with standardized labeling and reference-based classification.
    """

    # Add RAG context
    rag_context = retrieve_context(report_context)

    prompt = f"""
    You are a medical data extraction system.
    Analyze the following breast cancer pathology reports and extract biomarker information.

    Extract and return the following fields as a valid JSON object:
    - tumor_size (in cm)
    - her2 ("Negative", "Equivocal", "Positive", "Ordered, result pending", or "Not Reported")
    - er ("Positive", "Negative", or "Not Reported")
    - pr ("Positive", "Negative", or "Not Reported")
    - ki67 (percentage + category: Low / Borderline / High / Not Reported)
    - p53 ("Positive", "Negative", or "Not Reported")

    Reference Ranges (for classification only, do not include in output):
    - ER/PR: Negative <1%, Positive ≥1%
    - Ki-67: Low <10%, Borderline 10–20%, High >20%
    - HER2/neu: Negative 0–1+, Equivocal 2+, Positive 3+ (reflexed to FISH if equivocal)

    Output Format Example:
    {{
        "tumor_size": "1.0 cm",
        "her2": "Negative (0–1+)",
        "er": "Positive (98%, 3+)",
        "pr": "Positive (95%, 3+)",
        "ki67": "12% (Borderline)",
        "p53": "Not Reported"
    }}

    If a field is not found in the report, explicitly return "Not Reported".

    EXAMPLE REPORTS FOR CONTEXT:
    Report 1:
    Breast, left, simple mastectomy: Invasive ductal carcinoma, Nottingham grade III (of III), is identified forming a 6.0 x 4.0 x 3.5 cm tumor (AJCC pT3) in the central-upper inner region of the breast.
    Her-2/NEU has been ordered on paraffin-embedded tissue on the left breast.

    Expected output:
    {{
        "tumor_size": "6.0 x 4.0 x 3.5 cm (left)",
        "her2": "Ordered, result pending",
        "er": "Not Reported",
        "pr": "Not Reported",
        "ki67": "Not Reported",
        "p53": "Not Reported"
    }}

    Report 2:
    IHC-stainings: Estrogen receptor - positive reaction 98% 3+; Progesterone receptor - positive reaction 95% 3+;
    Her-2/neu - negative reaction; Ki67 - 12%.

    Expected output:
    {{
        "tumor_size": "1.0 cm (left)",
        "her2": "Negative (0–1+)",
        "er": "Positive (98%, 3+)",
        "pr": "Positive (95%, 3+)",
        "ki67": "12% (Borderline)",
        "p53": "Not Reported"
    }}

    RAG CONTEXT (for model background):
    {rag_context}

    CONTEXT TO ANALYZE:
    {report_context}
    """

    try:
        response = model.generate_content(prompt)
        text_output = response.text.strip()

        # Extract JSON safely
        try:
            result = json.loads(text_output)
        except json.JSONDecodeError:
            text_output = text_output[text_output.find("{"):text_output.rfind("}") + 1]
            result = json.loads(text_output)

        return (
            result.get("tumor_size", "Not Reported"),
            result.get("er", "Not Reported"),
            result.get("pr", "Not Reported"),
            result.get("her2", "Not Reported"),
            result.get("ki67", "Not Reported"),
            result.get("p53", "Not Reported"),
        )

    except Exception as e:
        print(f"Error with Gemini API: {e}")
        return ("Error", "Error", "Error", "Error", "Error", "Error")

# --- Main Execution ---
print("Initializing Gemini model...")
gemini_model = genai.GenerativeModel('gemini-2.5-flash-lite')

try:
    df = pd.read_excel(input_excel_file)
except FileNotFoundError:
    print(f"Error: '{input_excel_file}' not found.")
    exit()

filename_column = 'File_name'
report_text_column = 'ExtractedText'

if filename_column not in df.columns or report_text_column not in df.columns:
    print(f"Error: Excel file must contain '{filename_column}' and '{report_text_column}' columns.")
    exit()

print(f"Processing {len(df)} pathology reports...")

for index, row in df.iterrows():
    file_name = row[filename_column]
    report_text = row[report_text_column]

    if not isinstance(report_text, str) or len(report_text.strip()) < 10:
        writeInSheets(file_name, "Skipped", "Skipped", "Skipped", "Skipped", "Skipped", "Skipped")
        continue

    print(f"\n--- Processing Report: {file_name} ---")
    tumor_size, er, pr, her2, ki67, p53 = extract_features_with_gemini(report_text, gemini_model)

    print(f"Tumor Size: {tumor_size}")
    print(f"ER: {er}")
    print(f"PR: {pr}")
    print(f"HER2: {her2}")
    print(f"Ki-67: {ki67}")
    print(f"p53: {p53}")

    writeInSheets(file_name, tumor_size, er, pr, her2, ki67, p53)

print(f"\n✅ Extraction complete. Results saved to {output_excel_file}")
