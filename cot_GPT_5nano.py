!pip install openai pandas openpyxl --quiet

import os
import json
import pandas as pd
from openpyxl import Workbook, load_workbook
from openai import OpenAI
from google.colab import files
from tqdm import tqdm

OPENAI_API_KEY = ""
client = OpenAI(api_key=OPENAI_API_KEY)

input_excel_file = "Modified_data.xlsx"
output_excel_file = "features_openai.xlsx"

if not os.path.exists(input_excel_file):
    print(f"'{input_excel_file}' not found. Please upload it now.")
    uploaded = files.upload()
    if input_excel_file not in uploaded:
        print(f"Please upload a file named '{input_excel_file}'.")
        raise SystemExit
    print(f"Uploaded: {list(uploaded.keys())}")

def write_to_excel(file_name, tumor_size, er, pr, her2, ki67, p53):
    """Appends extracted features into an Excel file."""
    if not os.path.exists(output_excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Features"
        headers = ["Source File Name", "Tumor Size", "ER", "PR", "HER2", "Ki67", "p53"]
        ws.append(headers)
        wb.save(output_excel_file)

    wb = load_workbook(output_excel_file)
    ws = wb.active
    ws.append([file_name, tumor_size, er, pr, her2, ki67, p53])
    wb.save(output_excel_file)

def extract_features_openai(report_context, rag_context=""):
    """
    Uses a strict prompt that requests an external extraction_rationale (1-3 short bullets per biomarker).
    Returns a dict with keys:
      tumor_size, her2, er, pr, ki67, p53, extraction_rationale
    where extraction_rationale is itself a dict mapping biomarker -> list of short bullets.
    """

    prompt = f"""
You are a medical data extraction system. Do NOT reveal internal chain-of-thought or private model reasoning.
Instead, produce a concise, external step-by-step extraction log (1-3 short bullets per biomarker) called "extraction_rationale".

Task:
From the pathology report below, extract and return a JSON object with these keys:
- tumor_size (string; include units and laterality if present, or "Not Reported")
- her2 (one of: "Negative", "Equivocal", "Positive", "Ordered, result pending", or "Not Reported")
- er (one of: "Positive", "Negative", or "Not Reported"; if percent/score present include, e.g. "Positive (98%, 3+)")
- pr (same format as ER)
- ki67 (percentage + category: "12% (Borderline)" or "Not Reported")
- p53 (one of: "Positive", "Negative", "Not Reported")

Also include:
- extraction_rationale: a JSON object mapping each biomarker name to a short array of 1-3 succinct bullet strings that list the textual cues used to decide the value (e.g., ["ER: 'positive reaction 98% 3+' found", "Percent >=1% => Positive"]).

Reference rules for categorization (for model only; do not print these as part of rationale):
- ER/PR: Negative <1%, Positive ≥1%
- Ki-67: Low <10%, Borderline 10–20%, High >20%
- HER2: Negative 0–1+, Equivocal 2+, Positive 3+ (equivocal often reflexed to FISH)

If a field is not present in the report, set it to "Not Reported" and in rationale write a single bullet like "Not found".

Output format (strict JSON). Example:
{{
  "tumor_size": "6.0 x 4.0 x 3.5 cm (left)",
  "her2": "Ordered, result pending",
  "er": "Not Reported",
  "pr": "Not Reported",
  "ki67": "Not Reported",
  "p53": "Not Reported",
  "extraction_rationale": {{
      "tumor_size": ["Found 'forming a 6.0 x 4.0 x 3.5 cm tumor'"],
      "her2": ["Phrase 'Her-2/NEU has been ordered' -> Ordered, result pending"],
      "er": ["Not found"]
  }}
}}

RAG CONTEXT: {rag_context}

PATHOLOGY REPORT:
{report_context}

Return only valid JSON—no extra commentary.
"""

    try:
        response = client.chat.completions.create(
            model="gpt-5-nano",
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
        )

        text_output = response.choices[0].message.content.strip()

        try:
            parsed = json.loads(text_output)
        except json.JSONDecodeError:
            start = text_output.find("{")
            end = text_output.rfind("}")
            if start != -1 and end != -1 and end > start:
                json_text = text_output[start:end+1]
                parsed = json.loads(json_text)
            else:
                raise
        def ensure_field(d, key, default="Not Reported"):
            return d.get(key, default)

        tumor_size = ensure_field(parsed, "tumor_size", "Not Reported")
        her2 = ensure_field(parsed, "her2", "Not Reported")
        er = ensure_field(parsed, "er", "Not Reported")
        pr = ensure_field(parsed, "pr", "Not Reported")
        ki67 = ensure_field(parsed, "ki67", "Not Reported")
        p53 = ensure_field(parsed, "p53", "Not Reported")

        raw_rationale = parsed.get("extraction_rationale", {})
        extraction_rationale = {}
        for key in ["tumor_size", "her2", "er", "pr", "ki67", "p53"]:
            val = raw_rationale.get(key)
            if isinstance(val, list) and all(isinstance(x, str) for x in val) and len(val) > 0:
                extraction_rationale[key] = val[:3]
            else:
                if key in parsed and parsed[key] != "Not Reported":
                    extraction_rationale[key] = ["Found but no short rationale provided"]
                else:
                    extraction_rationale[key] = ["Not found"]

        result = {
            "tumor_size": tumor_size,
            "her2": her2,
            "er": er,
            "pr": pr,
            "ki67": ki67,
            "p53": p53,
            "extraction_rationale": extraction_rationale,
        }

        return result

    except Exception as e:
        print(f"Error while calling OpenAI API: {e}")
        return {
            "tumor_size": "Error",
            "her2": "Error",
            "er": "Error",
            "pr": "Error",
            "ki67": "Error",
            "p53": "Error",
            "extraction_rationale": {
                "tumor_size": [f"Error: {e}"],
                "her2": [f"Error: {e}"],
                "er": [f"Error: {e}"],
                "pr": [f"Error: {e}"],
                "ki67": [f"Error: {e}"],
                "p53": [f"Error: {e}"],
            },
        }

print("Initializing extraction...")

try:
    df = pd.read_excel(input_excel_file)
except Exception as e:
    print(f"Error reading Excel file: {e}")
    raise SystemExit

filename_col = "File_name"
text_col = "text"

if filename_col not in df.columns or text_col not in df.columns:
    print(f"Missing columns. Ensure '{filename_col}' and '{text_col}' exist.")
    raise SystemExit

print(f"Found {len(df)} reports to process.\n")

for idx, row in tqdm(df.iterrows(), total=len(df), desc="Extracting features"):
    file_name = row[filename_col]
    report_context = row[text_col]

    if not isinstance(report_context, str) or len(report_context) < 10:
        write_to_excel(file_name, "Skipped", "Skipped", "Skipped", "Skipped", "Skipped", "Skipped")
        continue

    result = extract_features_openai(report_context)
    write_to_excel(
        file_name,
        result["tumor_size"],
        result["er"],
        result["pr"],
        result["her2"],
        result["ki67"],
        result["p53"],
    )

print(f"\n Done! Results saved in: {output_excel_file}")
